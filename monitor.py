
import sys
import os
import openpyxl
import uuid
import time
import datetime

class Monitor:

    archivoMaestro : str;
    directorioDeposito : str;
    frecuenciaMonitor : str;

    def __init__(self, archivoMaestro:str, directorioDeposito:str, frecuenciaMonitor:int):
        self.archivoMaestro = archivoMaestro;
        self.directorioDeposito = directorioDeposito;
        self.frecuenciaMonitor =frecuenciaMonitor;
        print('Scan Started!');
        print('Master Excel File : ',self.archivoMaestro);
        print('Target Path       : ',self.directorioDeposito);
        print('Frecuency in ms   : ',self.frecuenciaMonitor);
        print('Press [CTRL][C] to stop scanning');

    def revisarCarpeta(self):
        item: str;
        itemPath : str;
        partesNombre : list;
        listaHojas : list;
        nombreHojaOrigen :str;
        dirCarpeta : list = os.listdir(self.directorioDeposito);
        libroDestino = openpyxl.load_workbook(self.archivoMaestro);
        for item in dirCarpeta:
            itemPath =  os.path.join(self.directorioDeposito,item);
            if os.path.isfile(itemPath):
                partesNombre = item.split('.');
                if(partesNombre[len(partesNombre)-1]=='xlsx'):
                    libroOrigen = openpyxl.load_workbook(itemPath);
                    listaHojas = libroOrigen.sheetnames;
                    for nombreHojaOrigen in listaHojas:
                        hojaOrigen = libroOrigen[nombreHojaOrigen];
                        hojaDestino = None;
                        try:
                            hojaDestino = libroDestino.create_sheet(title=nombreHojaOrigen);
                        except:
                            hojaDestino = libroDestino.create_sheet(title=nombreHojaOrigen+str(uuid.uuid4()));
                        for filaOrigen in hojaOrigen.iter_rows(min_row=1, max_row=hojaOrigen.max_row, values_only=True):
                            hojaDestino.append(filaOrigen);
                    libroDestino.save(self.archivoMaestro);
                    libroOrigen.close();
        libroDestino.close();
    
    def limpiarCarpeta(self):
        itemName: str;
        itemRename : str;
        itemPath : str;
        itemPathDest : str;
        dirCarpeta : list = os.listdir(self.directorioDeposito);
        dirProcesado :str = os.path.join(self.directorioDeposito,'Processed');
        dirNoProcesado :str = os.path.join(self.directorioDeposito,'Not applicable');
        partesNombre : list;
        if(not os.path.exists(dirNoProcesado)):
            os.makedirs(dirNoProcesado);
        if(not os.path.exists(dirProcesado)):
            os.makedirs(dirProcesado);
        for itemName in dirCarpeta:
            itemPath =  os.path.join(self.directorioDeposito,itemName);
            if os.path.isfile(itemPath):
                partesNombre = itemName.split('.');
                if(partesNombre[len(partesNombre)-1]=='xlsx'):
                    itemPathDest = os.path.join(dirProcesado,itemName);
                    if(os.path.exists(itemPathDest)):
                        itemRename = str(uuid.uuid4())+'.'+itemName;
                        itemPathDest = os.path.join(dirProcesado,itemRename);
                else:
                    itemPathDest = os.path.join(dirNoProcesado,itemName);
                    if(os.path.exists(itemPathDest)):
                        itemRename = str(uuid.uuid4())+'.'+itemName;
                        itemPathDest = os.path.join(dirNoProcesado,itemRename);
            else:
                itemPathDest = os.path.join(dirNoProcesado,itemName);
                if(os.path.exists(itemPathDest)):
                    itemRename = str(uuid.uuid4())+'.'+itemName;
                    itemPathDest = os.path.join(dirNoProcesado,itemRename);
            if(not(itemName=='Processed' or itemName=='Not applicable')):
                os.rename(itemPath,itemPathDest);

    def cicloMonitor(self):
        while(True):
            print('Scanned at: ',datetime.datetime.now());
            self.revisarCarpeta();
            self.limpiarCarpeta();
            time.sleep(self.frecuenciaMonitor/1000);

    
def main():
    if(len(sys.argv)>3):
        archivoMaestro : str = sys.argv[1];
        directorioDeposito : str = sys.argv[2];
        if(sys.argv[3].isdigit()):
            frecuenciaMonitor : int = int(sys.argv[3]);
        else:
            frecuenciaMonitor : int = 1000;
        if(os.path.isfile(archivoMaestro) and os.path.isdir(directorioDeposito)):
            nuevoMonitor : Monitor = Monitor(archivoMaestro,directorioDeposito,frecuenciaMonitor);
            try:
                nuevoMonitor.cicloMonitor();
            except KeyboardInterrupt:
                print('Scan stopped!');
        else:
            printModoDeUso();
    else:
        printModoDeUso();

def printModoDeUso():
    print('Usage mode:');
    print('python monitor.py <master-excel-file> <target-path> <frecuency-milliseconds>');
    print('     <master-excel-file>      :  File where information will be joined (.xlsx)');
    print('     <target-path>            :  Path to be scanned for new files');
    print('     <frecuency-milliseconds> :  Frecuency of monitoring defult=60000');


if __name__=="__main__":
    main();
